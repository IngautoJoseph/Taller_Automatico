
import streamlit as st
import pandas as pd
import psycopg2
from fpdf import FPDF
import openpyxl
import io
import os
import uuid
import smtplib
from email.message import EmailMessage
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill

# CONFIGURACIÓN
CORREO_ADMIN = "accesoriossd@ingauto.com.ec"
LOGO_URL = "https://www.ingauto.com.ec/wp-content/uploads/2019/06/logo-Ingauto-T.png"

def conectar_postgres():
    return psycopg2.connect(
        host="switchback.proxy.rlwy.net",
        port=25570,
        dbname="railway",
        user="postgres",
        password="ejyethsIreptiLCiEEITaElUdVliccKM"
    )

def inicializar_bd():
    conn = conectar_postgres()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS citas (
            id UUID PRIMARY KEY,
            numero_cita TEXT,
            nombre TEXT, telefono TEXT, cedula TEXT, correo TEXT,
            marca TEXT, modelo TEXT, anio TEXT, placa TEXT, kilometraje TEXT,
            combustible TEXT, motor TEXT, chasis TEXT,
            servicio TEXT, servicio_extra TEXT, fecha TEXT, hora TEXT
        );
    """)
    conn.commit()
    conn.close()

def obtener_numero_cita():
    conn = conectar_postgres()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM citas")
    count = cur.fetchone()[0] + 1
    conn.close()
    return f"CITA N.º {str(count).zfill(3)}"

def guardar_cita(datos):
    conn = conectar_postgres()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO citas (
            id, numero_cita, nombre, telefono, cedula, correo,
            marca, modelo, anio, placa, kilometraje, combustible,
            motor, chasis, servicio, servicio_extra, fecha, hora
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (
        str(uuid.uuid4()), datos["numero_cita"], datos["nombre"], datos["telefono"],
        datos["cedula"], datos["correo"], datos["marca"], datos["modelo"],
        datos["anio"], datos["placa"], datos["kilometraje"], datos["combustible"],
        datos["motor"], datos["chasis"], datos["servicio"], datos["servicio_extra"],
        datos["fecha"], datos["hora"]
    ))
    conn.commit()
    conn.close()

def generar_pdf(datos, archivo):
    pdf = FPDF()
    pdf.add_page()
    pdf.image(LOGO_URL, x=60, w=90)
    pdf.set_font("Arial", "B", 14)
    pdf.ln(35)
    pdf.cell(200, 10, txt=datos["numero_cita"], ln=True, align="C")
    pdf.cell(200, 10, txt="CITA REGISTRADA - INGAUTO", ln=True, align="C")
    pdf.ln(10)
    pdf.set_font("Arial", "", 12)
    for k, v in datos.items():
        if k != "numero_cita":
            pdf.cell(60, 10, txt=f"{k.upper()}:", ln=0)
            pdf.cell(120, 10, txt=str(v).upper(), ln=1)
    pdf.output(archivo)

def enviar_correo_pdf(destinatario, archivo_pdf):
    msg = EmailMessage()
    msg["Subject"] = "Cita registrada - Ingauto"
    msg["From"] = CORREO_ADMIN
    msg["To"] = [destinatario, CORREO_ADMIN]
    msg.set_content("Adjunto se encuentra el comprobante PDF de su cita registrada.")
    with open(archivo_pdf, "rb") as f:
        msg.add_attachment(f.read(), maintype="application", subtype="pdf", filename=os.path.basename(archivo_pdf))
    try:
        with smtplib.SMTP_SSL("mail.ingauto.com.ec", 465) as smtp:
            smtp.login(CORREO_ADMIN, "51TBdC375q")
            smtp.send_message(msg)
    except Exception as e:
        st.error(f"Error al enviar correo: {e}")

def exportar_excel():
    conn = conectar_postgres()
    df = pd.read_sql("SELECT * FROM citas ORDER BY fecha DESC", conn)
    df.insert(0, "#", range(1, len(df) + 1))
    conn.close()
    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Citas')
    ws = writer.book['Citas']
    header_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill(start_color="FF7300", end_color="FF7300", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    writer.save()
    return output.getvalue()

# === INTERFAZ STREAMLIT ===
st.set_page_config(page_title="Citas Ingauto", layout="wide")
st.image(LOGO_URL, width=500)
st.title("📋 Sistema de Citas - Ingauto")

inicializar_bd()

st.header("Registrar nueva cita")

with st.form("formulario"):
    numero_cita = obtener_numero_cita()
    st.markdown(f"### {numero_cita}")
    nombre = st.text_input("Nombre completo")
    telefono = st.text_input("Teléfono")
    cedula = st.text_input("Cédula / RUC")
    correo = st.text_input("Correo del cliente")
    marca = st.text_input("Marca")
    modelo = st.text_input("Modelo")
    anio = st.text_input("Año")
    placa = st.text_input("Placa")
    kilometraje = st.text_input("Kilometraje")
    combustible = st.selectbox("Combustible", ["Gasolina", "Diésel"])
    motor = st.text_input("Motor")
    chasis = st.text_input("Chasis")
    servicio = st.text_input("Servicio")
    servicio_extra = st.text_area("Servicio extra")
    fecha = st.date_input("Fecha")
    hora = st.time_input("Hora")
    enviar = st.form_submit_button("Registrar y generar PDF")

if enviar:
    datos = {
        "numero_cita": numero_cita,
        "nombre": nombre, "telefono": telefono, "cedula": cedula, "correo": correo,
        "marca": marca, "modelo": modelo, "anio": anio, "placa": placa,
        "kilometraje": kilometraje, "combustible": combustible, "motor": motor,
        "chasis": chasis, "servicio": servicio, "servicio_extra": servicio_extra,
        "fecha": str(fecha), "hora": str(hora)
    }
    archivo_pdf = f"/tmp/{numero_cita.replace(' ', '_')}_{placa}.pdf"
    generar_pdf(datos, archivo_pdf)
    guardar_cita(datos)
    enviar_correo_pdf(correo, archivo_pdf)
    st.success("✅ Cita registrada y correo enviado")
    with open(archivo_pdf, "rb") as f:
        st.download_button("⬇️ Descargar PDF", f, file_name=os.path.basename(archivo_pdf))

st.header("📑 Citas registradas")

try:
    conn = conectar_postgres()
    df = pd.read_sql("SELECT * FROM citas ORDER BY fecha DESC, hora DESC", conn)
    conn.close()
    if not df.empty:
        df.index = range(1, len(df) + 1)
        st.dataframe(df)
    else:
        st.info("No hay citas registradas aún.")
except Exception as e:
    st.error(f"Error al mostrar citas: {e}")

if st.button("📊 Exportar Excel bonito"):
    excel_bytes = exportar_excel()
    st.download_button("⬇️ Descargar Excel", data=excel_bytes, file_name="citas_ingauto.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
