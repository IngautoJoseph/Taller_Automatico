
import streamlit as st
from fpdf import FPDF
import pandas as pd
import psycopg2
import io
import os
import uuid
import smtplib
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from email.message import EmailMessage

# Configuración de conexión a PostgreSQL
def conectar_postgres():
    return psycopg2.connect(
        host="switchback.proxy.rlwy.net",
        port=25570,
        dbname="railway",
        user="postgres",
        password="ejyethsIreptiLCiEEITaElUdVliccKM"
    )

# Configuración del correo corporativo
SMTP_SERVER = "mail.ingauto.com.ec"
SMTP_PORT = 465
EMAIL_USER = "accesoriossd@ingauto.com.ec"
EMAIL_PASS = "51TBdC375q"

# Enviar PDF al correo del cliente
def enviar_correo_pdf(destinatario, ruta_pdf):
    try:
        msg = EmailMessage()
        msg["Subject"] = "Cita registrada - Ingauto"
        msg["From"] = EMAIL_USER
        msg["To"] = destinatario
        msg.set_content("Adjunto se encuentra el comprobante PDF de su cita en Ingauto Catamayo.")

        with open(ruta_pdf, "rb") as f:
            file_data = f.read()
            file_name = os.path.basename(ruta_pdf)

        msg.add_attachment(file_data, maintype="application", subtype="pdf", filename=file_name)

        with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT) as smtp:
            smtp.login(EMAIL_USER, EMAIL_PASS)
            smtp.send_message(msg)
        return True
    except Exception as e:
        st.error(f"Error al enviar el correo: {e}")
        return False

# Exportar a Excel con formato
def exportar_excel():
    try:
        conn = conectar_postgres()
        df = pd.read_sql("SELECT * FROM citas ORDER BY fecha DESC, hora DESC", conn)
        df.insert(0, "#", range(1, len(df) + 1))
        conn.close()

        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Citas"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="FF7300", end_color="FF7300", fill_type="solid")
        center = Alignment(horizontal="center")

        for col_num, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        for i, row in enumerate(df.values, 2):
            for j, val in enumerate(row, 1):
                ws.cell(row=i, column=j, value=val)

        wb.save(output)
        return output.getvalue()
    except Exception as e:
        st.error(f"Error al exportar Excel: {e}")
        return None

# Crear tabla si no existe
def inicializar_bd():
    conn = conectar_postgres()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS citas (
            id UUID PRIMARY KEY,
            nombre TEXT, telefono TEXT, cedula TEXT, correo TEXT,
            marca TEXT, modelo TEXT, anio TEXT, placa TEXT, kilometraje TEXT,
            combustible TEXT, motor TEXT, chasis TEXT,
            servicio TEXT, servicio_extra TEXT, fecha TEXT, hora TEXT
        )
    """)
    conn.commit()
    conn.close()

# Guardar cita
def guardar_cita(datos):
    try:
        conn = conectar_postgres()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO citas (
                id, nombre, telefono, cedula, correo, marca, modelo, anio, placa, kilometraje,
                combustible, motor, chasis, servicio, servicio_extra, fecha, hora
            ) VALUES (
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s, %s, %s
            )
        """, (str(uuid.uuid4()), *datos.values()))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        st.error(f"Error al guardar cita: {e}")
        return False

# Generar PDF
direccion_logo = "https://www.ingauto.com.ec/wp-content/uploads/2019/06/logo-Ingauto-T.png"
def generar_pdf(datos, ruta):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=14)
    pdf.image(direccion_logo, x=10, y=8, w=80)
    pdf.ln(30)
    pdf.cell(200, 10, txt="Cita Registrada - PostgreSQL", ln=True, align="C")
    pdf.ln(10)
    pdf.set_fill_color(255, 115, 0)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Arial", 'B', 12)
    pdf.cell(60, 10, "Campo", 1, 0, 'C', fill=True)
    pdf.cell(130, 10, "Valor", 1, 1, 'C', fill=True)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Arial", size=12)
    for clave, valor in datos.items():
        pdf.cell(60, 10, clave, 1)
        pdf.cell(130, 10, str(valor), 1)
        pdf.ln()
    pdf.output(ruta)

# INTERFAZ STREAMLIT
st.image(direccion_logo, width=600)
st.title("Sistema de Citas - PostgreSQL + Excel + Email")

with st.form("formulario_cita"):
    nombre = st.text_input("Nombre completo")
    telefono = st.text_input("Teléfono")
    cedula = st.text_input("Cédula / RUC")
    correo_cliente = st.text_input("Correo del cliente")
    marca = st.text_input("Marca del vehículo")
    modelo = st.text_input("Modelo")
    anio = st.text_input("Año")
    placa = st.text_input("Placa")
    kilometraje = st.text_input("Kilometraje")
    combustible = st.selectbox("Combustible", ["Gasolina", "Diésel"])
    motor = st.text_input("Motor")
    chasis = st.text_input("Chasis")
    servicio = st.selectbox("Servicio solicitado", [
        "Mantenimiento 5000 km", "Mantenimiento 10000 km",
        "Cambio de aceite", "Revisión de frenos", "Diagnóstico general", "Otros"
    ])
    servicio_extra = st.text_area("📝 Descripción del servicio solicitado (si aplica)")
    fecha = st.date_input("Fecha de cita")
    hora = st.time_input("Hora")
    enviar = st.form_submit_button("Registrar y generar PDF")

inicializar_bd()

if enviar:
    if not all([nombre, telefono, cedula, correo_cliente, marca, modelo, anio, placa]):
        st.warning("Por favor completa todos los campos obligatorios.")
    else:
        datos = {
            "nombre": nombre,
            "telefono": telefono,
            "cedula": cedula,
            "correo": correo_cliente,
            "marca": marca,
            "modelo": modelo,
            "anio": anio,
            "placa": placa,
            "kilometraje": kilometraje,
            "combustible": combustible,
            "motor": motor,
            "chasis": chasis,
            "servicio": servicio,
            "servicio_extra": servicio_extra,
            "fecha": str(fecha),
            "hora": str(hora)
        }
        nombre_pdf = f"cita_{placa}_{fecha}.pdf"
        ruta_pdf = os.path.join("/tmp", nombre_pdf)
        generar_pdf(datos, ruta_pdf)
        if guardar_cita(datos):
            st.success("✅ Cita guardada en PostgreSQL.")
            if enviar_correo_pdf(correo_cliente, ruta_pdf):
                st.success(f"📩 PDF enviado a {correo_cliente}")
        with open(ruta_pdf, "rb") as f:
            st.download_button("⬇️ Descargar PDF", data=f, file_name=nombre_pdf, mime="application/pdf")

# Mostrar citas
st.subheader("📋 Citas registradas")
try:
    conn = conectar_postgres()
    df = pd.read_sql("SELECT * FROM citas ORDER BY fecha DESC, hora DESC", conn)
    df.insert(0, "#", range(1, len(df) + 1))
    st.dataframe(df)
    conn.close()
except Exception as e:
    st.error(f"Error al cargar citas: {e}")

# Exportar Excel
st.subheader("⬇️ Exportar citas")
if st.button("Generar y descargar Excel"):
    excel_bytes = exportar_excel()
    if excel_bytes:
        st.download_button(
            label="📊 Descargar archivo Excel",
            data=excel_bytes,
            file_name="citas_postgres.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
