

import streamlit as st
import pandas as pd
import psycopg2
from fpdf import FPDF
import io
import os
import uuid
import smtplib
from email.message import EmailMessage
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# ====================== CONFIGURACIÓN ======================
LOGO_URL = "https://www.ingauto.com.ec/wp-content/uploads/2019/06/logo-Ingauto-T.png"
CORREO_ADMIN = "accesoriossd@ingauto.com.ec"
PASS_CORREO = "51TBdC375q"
SMTP_SERVER = "mail.ingauto.com.ec"
SMTP_PORT = 465

# ====================== CONEXIÓN BD ======================
def conectar_postgres():
    return psycopg2.connect(
        host="switchback.proxy.rlwy.net",
        port=25570,
        dbname="railway",
        user="postgres",
        password="ejyethsIreptiLCiEEITaElUdVliccKM"
    )

# ====================== UTILIDADES ======================
def inicializar_bd():
    conn = conectar_postgres()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS citas (
            id UUID PRIMARY KEY,
            numero_cita TEXT,
            nombre TEXT, telefono TEXT, cedula TEXT, correo TEXT,
            marca TEXT, modelo TEXT, anio TEXT, placa TEXT, kilometraje TEXT,
            combustible TEXT, motor TEXT, chasis TEXT,
            servicio TEXT, servicio_extra TEXT, fecha TEXT, hora TEXT
        );
    """)
    conn.commit()
    conn.close()

def obtener_numero_cita():
    conn = conectar_postgres()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM citas")
    count = cur.fetchone()[0] + 1
    conn.close()
    return f"CITA N.º {str(count).zfill(3)}"

def guardar_cita(datos, editar=False, cita_id=None):
    conn = conectar_postgres()
    cur = conn.cursor()
    if editar and cita_id:
        cur.execute("""
            UPDATE citas SET
                nombre=%s, telefono=%s, cedula=%s, correo=%s,
                marca=%s, modelo=%s, anio=%s, placa=%s, kilometraje=%s,
                combustible=%s, motor=%s, chasis=%s, servicio=%s,
                servicio_extra=%s, fecha=%s, hora=%s
            WHERE id=%s
        """, (
            datos["nombre"], datos["telefono"], datos["cedula"], datos["correo"],
            datos["marca"], datos["modelo"], datos["anio"], datos["placa"], datos["kilometraje"],
            datos["combustible"], datos["motor"], datos["chasis"], datos["servicio"],
            datos["servicio_extra"], datos["fecha"], datos["hora"], cita_id
        ))
    else:
        cur.execute("""
            INSERT INTO citas (
                id, numero_cita, nombre, telefono, cedula, correo,
                marca, modelo, anio, placa, kilometraje, combustible,
                motor, chasis, servicio, servicio_extra, fecha, hora
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            str(uuid.uuid4()), datos["numero_cita"], datos["nombre"], datos["telefono"],
            datos["cedula"], datos["correo"], datos["marca"], datos["modelo"],
            datos["anio"], datos["placa"], datos["kilometraje"], datos["combustible"],
            datos["motor"], datos["chasis"], datos["servicio"], datos["servicio_extra"],
            datos["fecha"], datos["hora"]
        ))
    conn.commit()
    conn.close()



def generar_pdf(datos, archivo):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.set_margins(15, 15, 15)

    # Logo centrado
    pdf.image(LOGO_URL, x=55, w=100)
    pdf.set_font("Arial", "B", 14)
    pdf.ln(35)
    pdf.cell(0, 10, txt=datos["numero_cita"], ln=True, align="C")
    pdf.cell(0, 10, txt="CITA REGISTRADA - INGAUTO", ln=True, align="C")
    pdf.ln(10)

    etiquetas = {
        "nombre": "Nombre",
        "telefono": "Teléfono",
        "cedula": "Cédula / RUC",
        "correo": "Correo",
        "marca": "Marca",
        "modelo": "Modelo",
        "anio": "Año",
        "placa": "Placa",
        "kilometraje": "Kilometraje",
        "combustible": "Combustible",
        "motor": "Motor",
        "chasis": "Chasis",
        "servicio": "Servicio",
        "servicio_extra": "Servicio adicional",
        "fecha": "Fecha",
        "hora": "Hora"
    }

    pdf.set_font("Arial", size=11)

    for clave, etiqueta in etiquetas.items():
        valor = str(datos.get(clave, "")).upper()
        pdf.set_fill_color(240, 240, 240)
        pdf.set_font("Arial", "B", 11)
        pdf.cell(60, 8, f"{etiqueta}:", border=1, fill=True)
        pdf.set_font("Arial", "", 11)
        pdf.cell(0, 8, valor, ln=True, border=1)

    pdf.output(archivo)



def enviar_correo_pdf(destinatario, archivo_pdf):
    msg = EmailMessage()
    msg["Subject"] = "Cita registrada - Ingauto"
    msg["From"] = CORREO_ADMIN
    msg["To"] = [destinatario, CORREO_ADMIN]
    msg.set_content("Adjunto se encuentra el comprobante PDF de su cita.")
    with open(archivo_pdf, "rb") as f:
        msg.add_attachment(f.read(), maintype="application", subtype="pdf", filename=os.path.basename(archivo_pdf))
    with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT) as smtp:
        smtp.login(CORREO_ADMIN, PASS_CORREO)
        smtp.send_message(msg)

def exportar_excel():
    conn = conectar_postgres()
    df = pd.read_sql("SELECT * FROM citas ORDER BY fecha DESC", conn)
    df.insert(0, "#", range(1, len(df)+1))
    conn.close()
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Citas"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF7300", end_color="FF7300", fill_type="solid")
    for col_num, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for i, row in enumerate(df.values, 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
    wb.save(output)
    return output.getvalue()

# ====================== APP ======================
st.set_page_config(page_title="Citas Ingauto", layout="wide")
st.image(LOGO_URL, width=500)
st.title("📋 Sistema de Citas - Ingauto")
inicializar_bd()

# --- FORMULARIO ---
if "editar_id" not in st.session_state:
    st.session_state.editar_id = None

st.subheader("Registrar / Editar Cita")

with st.form("formulario"):
    editar = st.session_state.editar_id is not None
    numero_cita = obtener_numero_cita() if not editar else st.session_state.numero_cita
    if editar:
        st.markdown(f"**Editando {numero_cita}**")
    else:
        st.markdown(f"**{numero_cita}**")

    nombre = st.text_input("Nombre completo")
    telefono = st.text_input("Teléfono")
    cedula = st.text_input("Cédula / RUC")
    correo = st.text_input("Correo del cliente")
    marca = st.text_input("Marca")
    modelo = st.text_input("Modelo")
    anio = st.text_input("Año")
    placa = st.text_input("Placa")
    kilometraje = st.text_input("Kilometraje")
    combustible = st.selectbox("Combustible", ["Gasolina", "Diésel"])
    motor = st.text_input("Motor")
    chasis = st.text_input("Chasis")
    servicio = st.selectbox("Servicio", ['MANTENIMIENTO 700 KM', 'MANTENIMIENTO 5,000 KM', 'MANTENIMIENTO 10,000 KM', 'MANTENIMIENTO 15,000 KM', 'MANTENIMIENTO 20,000 KM', 'MANTENIMIENTO 25,000 KM', 'MANTENIMIENTO 30,000 KM', 'MANTENIMIENTO 35,000 KM', 'MANTENIMIENTO 40,000 KM', 'MANTENIMIENTO 45,000 KM', 'MANTENIMIENTO 50,000 KM', 'CAMBIO DE ACEITE', 'ALINEACIÓN Y BALANCEO', 'DIAGNÓSTICO', 'OTRO'])
    servicio_extra = st.text_area("Servicio adicional")
    fecha = st.date_input("Fecha")
    hora = st.time_input("Hora")
    enviar = st.form_submit_button("Guardar")

    if enviar:
        campos = [nombre, telefono, cedula, correo, marca, modelo, anio, placa, kilometraje, motor, chasis, servicio]
        if any(c == "" for c in campos):
            st.warning("⚠️ Por favor, completa todos los campos obligatorios.")
        else:
            datos = {
                "numero_cita": numero_cita, "nombre": nombre, "telefono": telefono, "cedula": cedula, "correo": correo,
                "marca": marca, "modelo": modelo, "anio": anio, "placa": placa, "kilometraje": kilometraje,
                "combustible": combustible, "motor": motor, "chasis": chasis,
                "servicio": servicio, "servicio_extra": servicio_extra,
                "fecha": str(fecha), "hora": str(hora)
            }
            archivo_pdf = f"/tmp/{numero_cita.replace(' ', '_')}_{placa}.pdf"
            generar_pdf(datos, archivo_pdf)
            guardar_cita(datos, editar, st.session_state.editar_id)
            enviar_correo_pdf(correo, archivo_pdf)
            st.success("✅ Cita registrada y correo enviado")
st.session_state.editar_id = None

# --- TABLA ---
st.subheader("📑 Citas registradas")

conn = conectar_postgres()
df = pd.read_sql("SELECT * FROM citas ORDER BY fecha DESC, hora DESC", conn)
conn.close()

if not df.empty:
    for i, row in df.iterrows():
        with st.expander(f"{row['numero_cita']} - {row['nombre']} - {row['fecha']} {row['hora']}"):
            st.write(row.to_frame())
            col1, col2 = st.columns(2)
            with col1:
                if st.button(f"✏️ Editar {row['numero_cita']}", key=f"edit_{i}"):
                    for campo in df.columns:
                        if campo != "id":
                            st.session_state[campo] = row[campo]
                    st.session_state.numero_cita = row["numero_cita"]
                    st.session_state.editar_id = row["id"]
            with col2:
                ruta_pdf = f"/tmp/{row['numero_cita'].replace(' ', '_')}_{row['placa']}.pdf"
                generar_pdf(row, ruta_pdf)
                with open(ruta_pdf, "rb") as f:
                    st.download_button(f"📄 Descargar PDF {row['numero_cita']}", f, file_name=os.path.basename(ruta_pdf), key=f"pdf_{i}")
else:
    st.info("No hay citas registradas.")

if st.button("📊 Exportar Excel bonito"):
    excel = exportar_excel()
    st.download_button("⬇️ Descargar Excel", data=excel, file_name="citas_ingauto.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# Mostrar botón de descarga fuera del formulario
if 'archivo_pdf' in locals():
    with open(archivo_pdf, "rb") as f:
        st.download_button("⬇️ Descargar PDF", f, file_name=os.path.basename(archivo_pdf))
